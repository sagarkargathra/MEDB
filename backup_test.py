import pandas as pd
import glob
import openpyxl
import re

# Function to check if a row is a table note or footnote
def is_note_or_footnote(name):
    name_str = str(name)
    table_note_pattern = r'^\d+:'
    footnote_pattern = r'^[a-z]\)|#\)|\*\)'

    if re.match(table_note_pattern, name_str):
        return "Table notes"
    elif re.match(footnote_pattern, name_str):
        return "Footnotes"
    elif name_str.strip().lower() in ["footnotes", "notes"]:
        return name_str.strip().capitalize()
    return None

# Function to extract the 'Notes' section and return it as a DataFrame
def extract_notes(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    notes_data = []
    notes_started = False

    for row in sheet.iter_rows(values_only=True):
        if notes_started:
            notes_data.append(row)
        elif row[0] and str(row[0]).strip().lower() == "notes":
            notes_started = True
            notes_data.append(row)

    notes_df = pd.DataFrame(notes_data)
    return notes_df

# Function to unmerge cells and create a dictionary from footnotes
def extract_footnotes(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    footnotes_dict = {}
    
    for row in range(2, sheet.max_row + 1):
        footnote_key = sheet.cell(row=row, column=1).value
        if footnote_key:
            footnote_value = sheet.cell(row=row, column=2).value
            if footnote_value:
                footnotes_dict[footnote_key.strip()] = footnote_value.strip()
        for merged_cells in list(sheet.merged_cells.ranges):
            if merged_cells.min_row == row and merged_cells.min_col == 2:
                sheet.unmerge_cells(str(merged_cells))
    
    workbook.save(file_path)
    return footnotes_dict

# Function to process and read Excel files
def read_and_process_excel(file_path):
    footnotes_dict = extract_footnotes(file_path)
    df_full = pd.read_excel(file_path, header=None)
    df = pd.read_excel(file_path, header=1)
    df.columns = df.columns.map(str)
    df.set_index(df.columns[0], inplace=True)

    def replace_reference(reference):
        if pd.isna(reference):
            return reference
        matches = re.findall(r'([a-zA-Z]\)|\*\)|#\))', str(reference))
        replaced_reference = ", ".join([footnotes_dict.get(match.strip(')'), match) for match in matches])
        return replaced_reference.strip()

    if 'Reference' in df.columns:
        df['Reference'] = df['Reference'].apply(replace_reference)

    # Remove rows that are identified as "Table notes", "Footnotes", or "Notes"
    df = df[~df.index.to_series().apply(is_note_or_footnote).notnull()]

    # Ensure no "Table notes" remains in the Country column
    df = df[~df.index.str.lower().str.contains('table notes')]

    country = []
    year = []
    quantity = []
    reference = []
    unit = []
    sub_commodity = []

    unit_value = df_full.iloc[0, 1].strip()  # Remove leading/trailing whitespace from unit value

    for col in df.columns:
        if re.match(r'^\d{4}$', col):
            country.extend(df.index)
            year.extend([col] * len(df))
            quantity_col = col + ".1"
            quantity.extend(df[quantity_col] if quantity_col in df.columns else ["NA"] * len(df))
            reference.extend(df[col])
            unit.extend([unit_value] * len(df))
            sub_commodity.extend(df['Sub-commodity'] if 'Sub-commodity' in df.columns else ["NA"] * len(df))

    # Extract mineral name from cell A1 and remove the prefix "Production of "
    mineral_name = df_full.iloc[0, 0].replace("Production of ", "")

    processed_df = pd.DataFrame({
        'Mineral': mineral_name,
        'Sub-commodity': sub_commodity,
        'Country': country,
        'Year': year,
        'Quantity': quantity,
        'Unit': unit,
        'Reference': reference
    })

    processed_df.reset_index(drop=True, inplace=True)
    processed_df['Quantity'].replace('', 'NA', inplace=True)
    processed_df['Quantity'].fillna('NA', inplace=True)
    
    return processed_df

# Function to organize the notes and footnotes into the required format
def organize_notes(notes_df):
    table_notes_df = notes_df[notes_df.index.to_series().apply(is_note_or_footnote) == "Table notes"].drop_duplicates()
    footnotes_df = notes_df[notes_df.index.to_series().apply(is_note_or_footnote) == "Footnotes"].drop_duplicates()
    additional_notes_df = notes_df[notes_df.index.to_series().apply(is_note_or_footnote) == "Notes"].drop_duplicates()

    footnotes_df = footnotes_df.sort_index()

    organized_notes = pd.DataFrame()
    if not additional_notes_df.empty:
        notes_section = pd.DataFrame({"Table Notes": ["Notes"]})
        organized_notes = pd.concat([organized_notes, notes_section])
        additional_notes_df.insert(0, "Table Notes", additional_notes_df.iloc[:, 0])
        organized_notes = pd.concat([organized_notes, additional_notes_df.iloc[:, :1]])

    if not table_notes_df.empty:
        table_notes_section = pd.DataFrame({"Table Notes": ["Table notes"]})
        organized_notes = pd.concat([organized_notes, table_notes_section])
        table_notes_df.insert(0, "Table Notes", table_notes_df.iloc[:, 0])
        organized_notes = pd.concat([organized_notes, table_notes_df.iloc[:, :1]])

    if not footnotes_df.empty:
        footnotes_section = pd.DataFrame({"Table Notes": ["Footnotes"]})
        organized_notes = pd.concat([organized_notes, footnotes_section])
        footnotes_df.insert(0, "Table Notes", footnotes_df.iloc[:, 0])
        organized_notes = pd.concat([organized_notes, footnotes_df.iloc[:, :1]])

    return organized_notes

# Main function to process files, extract notes, and merge everything into the final output
def process_files_and_merge_notes():
    excel_files = glob.glob('files/*.xlsx')
    merged_df = pd.DataFrame()
    combined_notes = []

    for file in excel_files:
        df = read_and_process_excel(file)
        merged_df = pd.concat([merged_df, df])

        notes_df = extract_notes(file)
        for _, row in notes_df.iterrows():
            note_str = " ".join([str(x) for x in row if x is not None])
            combined_notes.append(note_str)

    combined_notes = list(set(combined_notes))

    filtered_notes = [note for note in combined_notes if re.match(r'^\d+:', note)]
    filtered_notes.sort(key=lambda x: int(re.match(r'(\d+):', x).group(1)))

    note_rows = pd.DataFrame({"Table Notes": filtered_notes})
    note_rows["Country"] = ""
    note_rows["Year"] = ""
    note_rows["Quantity"] = ""
    note_rows["Unit"] = ""
    note_rows["Reference"] = ""
    note_rows["Mineral"] = ""
    note_rows["Sub-commodity"] = ""

    merged_df = pd.concat([merged_df, note_rows], ignore_index=True)

    merged_df = merged_df.sort_values(by=['Country', 'Year'], na_position='last')

    extracted_star_note = next((note.split('*)')[1].strip() for note in combined_notes if isinstance(note, str) and note.startswith('*)')), None)
    extracted_hash_note = next((note.split('#)')[1].strip() for note in combined_notes if isinstance(note, str) and note.startswith('#)')), None)
    letter_notes = {letter: next((note.split(f'{letter})')[1].strip() for note in combined_notes if isinstance(note, str) and note.startswith(f'{letter})')), None) for letter in 'abcdefghijklmnopqrstuvwxyz'}

    def replace_references(row):
        if isinstance(row['Reference'], str):
            replacements = []

            if '*' in row['Reference'] and extracted_star_note:
                replacements.append(extracted_star_note)

            if '#' in row['Reference'] and extracted_hash_note:
                replacements.append(extracted_hash_note)

            for letter in 'abcdefghijklmnopqrstuvwxyz':
                if f'({letter})' in row['Reference'] and letter_notes[letter]:
                    replacements.append(letter_notes[letter])

            if replacements:
                row['Reference'] = ', '.join(replacements)

        return row

    merged_df = merged_df.apply(replace_references, axis=1)

    merged_df['Unit'] = merged_df['Unit'].apply(lambda x: "tonnes (metric)" if isinstance(x, str) and x.strip() == "" else x)

    # Replace all types of whitespace characters, including non-breaking spaces
    merged_df['Unit'] = merged_df['Unit'].apply(lambda x: re.sub(r'\s+', ' ', x).strip() if isinstance(x, str) else x)

    # Convert all values in the 'Quantity' column to strings to ensure left alignment
    merged_df['Quantity'] = merged_df['Quantity'].apply(lambda x: str(int(float(x))) if isinstance(x, (int, float)) else x)

     # Reorder columns to ensure 'Mineral' and 'Sub-commodity' are first
    columns_order = ['Mineral', 'Sub-commodity', 'Country', 'Year', 'Quantity', 'Unit', 'Reference', 'Table Notes']
    merged_df = merged_df[columns_order]

    output_file = 'output/final.xlsx'
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        merged_df.to_excel(writer, sheet_name='Data', index=False)

    final_df = pd.read_excel(output_file, sheet_name='Data')
    pd.set_option('display.max_columns', None)
    pd.set_option('display.max_rows', None)
    print(final_df.head())
    print(f'Final file saved to {output_file}')

# Run the process
process_files_and_merge_notes()
